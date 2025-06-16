# # main.py
#
# import os
# import time
# import threading
# import smtplib
# import winsound
# import cv2
# import numpy as np
# import openpyxl
# from datetime import datetime
# from openpyxl.styles import Font, Alignment
# from selenium import webdriver
# from selenium.webdriver.chrome.options import Options
# from selenium.webdriver.common.by import By
# from selenium.webdriver import Keys
# from selenium.common.exceptions import WebDriverException
# from email.mime.text import MIMEText
# from email.mime.multipart import MIMEMultipart
# from email.mime.base import MIMEBase
# from email import encoders
# from mss import mss
# from ctypes import POINTER, cast
# import comtypes
# from comtypes import CLSCTX_ALL
# from pycaw.pycaw import AudioUtilities, IAudioMeterInformation
#
# # -------------------- CONFIGURATION --------------------
# SCREEN_REGION = {"top": 200, "left": 200, "right": 200, "width": 1920, "height": 1200}
# ALERT_SOUND = "alert.mp3"
# EMAIL_ALERT = True
# EMAIL_FROM = "expressdartt@gmail.com"
# EMAIL_TO = "alibasti2021@gmail.com"
# EMAIL_CC = "m.minds456@gmail.com"
# SMTP_SERVER = "smtp.gmail.com"
# SMTP_PORT = 587
# SMTP_USERNAME = "expressdartt@gmail.com"
# SMTP_PASSWORD = "ztonyoyowxsdsarc"
# ALERT_COOLDOWN_SECONDS = 300
#
#
# # -------------------- UTILITIES --------------------
# def play_alert():
#     try:
#         print("[AUDIO] Playing beep alert...")
#         winsound.Beep(1000, 500)
#     except Exception as e:
#         print(f"[AUDIO] Beep failed: {e}")
#
#
# def send_email(subject, body, attachment_path=None):
#     if not EMAIL_ALERT:
#         return
#     try:
#         msg = MIMEMultipart()
#         msg["From"] = EMAIL_FROM
#         msg["To"] = EMAIL_TO
#         msg["CC"] = EMAIL_CC
#         msg["Subject"] = subject
#         msg.attach(MIMEText(body, "plain"))
#
#         if attachment_path:
#             with open(attachment_path, "rb") as f:
#                 part = MIMEBase("application", "octet-stream")
#                 part.set_payload(f.read())
#                 encoders.encode_base64(part)
#                 part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
#                 msg.attach(part)
#
#         recipients = [EMAIL_TO] + EMAIL_CC.split(",")
#         with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
#             server.starttls()
#             server.login(SMTP_USERNAME, SMTP_PASSWORD)
#             server.sendmail(EMAIL_FROM, recipients, msg.as_string())
#         print("[EMAIL] Alert sent.")
#     except Exception as e:
#         print(f"[EMAIL] Failed to send alert: {e}")
#
#
# def initialize_excel_log():
#     filename = f"freeze_log_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
#     if not os.path.exists(filename):
#         wb = openpyxl.Workbook()
#         ws = wb.active
#         ws.title = "Freeze Events"
#         ws.append(["Timestamp", "Event Description"])
#         for cell in ws[1]:
#             cell.font = Font(bold=True)
#             cell.alignment = Alignment(horizontal="center")
#         wb.save(filename)
#     return filename
#
#
# def log_freeze_event(message, excel_file):
#     timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#     print(f"[{timestamp}] {message}")
#     wb = openpyxl.load_workbook(excel_file)
#     ws = wb.active
#     ws.append([timestamp, message])
#     wb.save(excel_file)
#
#
# def is_stream_frozen(prev_frame, curr_frame, threshold=500):
#     diff = cv2.absdiff(prev_frame, curr_frame)
#     gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
#     return cv2.countNonZero(gray) < threshold
#
#
# def setup_browser(tester_url, stream_url):
#     options = Options()
#     options.add_argument("--start-maximized")
#     driver = webdriver.Chrome(options=options)
#     driver.get(tester_url)
#     time.sleep(3)
#     try:
#         input_field = driver.find_element(By.XPATH, "//*[@id='stream-url']")
#         driver.execute_script("arguments[0].scrollIntoView(true);", input_field)
#         time.sleep(1)
#         input_field.clear()
#         input_field.send_keys(stream_url)
#         input_field.send_keys(Keys.ENTER)
#         driver.find_element(By.XPATH, "//*[@id='rmp']/button").click()
#     except Exception as e:
#         print(f"[ERROR] Failed to enter stream URL: {e}")
#     time.sleep(8)
#     return driver
#
#
# # -------------------- MAIN MONITORING FUNCTION --------------------
# def main(tester_url, stream_url, stop_event):
#     excel_file = initialize_excel_log()
#
#     last_audio_alert_time = datetime.min
#     last_video_alert_time = datetime.min
#     audio_alert_suppressed = False
#     video_alert_suppressed = False
#
#     def should_send_audio_alert():
#         nonlocal last_audio_alert_time, audio_alert_suppressed
#         now = datetime.now()
#         if (now - last_audio_alert_time).total_seconds() >= ALERT_COOLDOWN_SECONDS:
#             last_audio_alert_time = now
#             audio_alert_suppressed = False
#             return True
#         elif not audio_alert_suppressed:
#             remaining = ALERT_COOLDOWN_SECONDS - (now - last_audio_alert_time).total_seconds()
#             print(f"[AUDIO ALERT PAUSED] Next audio alert in {int(remaining)} seconds.")
#             audio_alert_suppressed = True
#         return False
#
#     def should_send_video_alert():
#         nonlocal last_video_alert_time, video_alert_suppressed
#         now = datetime.now()
#         if (now - last_video_alert_time).total_seconds() >= ALERT_COOLDOWN_SECONDS:
#             last_video_alert_time = now
#             video_alert_suppressed = False
#             return True
#         elif not video_alert_suppressed:
#             remaining = ALERT_COOLDOWN_SECONDS - (now - last_video_alert_time).total_seconds()
#             print(f"[VIDEO ALERT PAUSED] Next video alert in {int(remaining)} seconds.")
#             video_alert_suppressed = True
#         return False
#
#     driver = setup_browser(tester_url, stream_url)
#
#     # -------------------- AUDIO MONITOR THREAD --------------------
#     def audio_thread():
#         comtypes.CoInitialize()
#         try:
#             threshold = 0.02
#             silence_duration = 2
#             silent_start = None
#             alerted = False
#             devices = AudioUtilities.GetSpeakers()
#             interface = devices.Activate(IAudioMeterInformation._iid_, CLSCTX_ALL, None)
#             meter = cast(interface, POINTER(IAudioMeterInformation))
#
#             while not stop_event.is_set():
#                 volume = meter.GetPeakValue()
#                 if volume < threshold:
#                     if silent_start is None:
#                         silent_start = time.time()
#                     elif time.time() - silent_start >= silence_duration and not alerted:
#                         if should_send_audio_alert():
#                             send_email("🔇 Audio Muted Alert", "Audio has been muted for more than 2 seconds.")
#                         alerted = True
#                 else:
#                     silent_start = None
#                     alerted = False
#                 time.sleep(0.1)
#         finally:
#             comtypes.CoUninitialize()
#
#     # -------------------- VIDEO MONITOR THREAD --------------------
#     def video_thread():
#         sct = mss()
#         prev_frame = None
#         freeze_count = 0
#         freeze_logged = False
#         freeze_email_sent = False
#         freeze_start_time = None
#
#         print("[INFO] Monitoring stream...")
#
#         while not stop_event.is_set():
#             img = sct.grab(SCREEN_REGION)
#             full_res_frame = np.array(img)
#             full_res_frame = cv2.cvtColor(full_res_frame, cv2.COLOR_BGRA2BGR)
#             frame = cv2.resize(full_res_frame, (1200, 720))
#
#             if prev_frame is not None:
#                 if is_stream_frozen(prev_frame, frame):
#                     freeze_count += 1
#
#                     if not freeze_logged:
#                         freeze_start_time = time.time()
#                         log_freeze_event("Freeze detected...", excel_file)
#                         freeze_logged = True
#                         freeze_email_sent = False
#
#                     if freeze_count >= 1:
#                         if should_send_video_alert():
#                             log_freeze_event("[ALERT] Stream appears to be frozen!", excel_file)
#                             freeze_image_path = "freeze_screenshot.jpg"
#                             cv2.imwrite(freeze_image_path, full_res_frame, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
#                             threading.Thread(target=play_alert, daemon=True).start()
#                             send_email("🚫 Stream Freeze Detected", "The stream appears frozen.", freeze_image_path)
#                             freeze_email_sent = True
#                         freeze_count = 0
#                 else:
#                     if freeze_logged:
#                         freeze_end_time = time.time()
#                         freeze_duration = freeze_end_time - freeze_start_time
#                         freeze_duration_str = time.strftime('%H:%M:%S', time.gmtime(freeze_duration))
#                         freeze_start_str = datetime.fromtimestamp(freeze_start_time).strftime('%Y-%m-%d %H:%M:%S')
#                         recovery_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#
#                         if freeze_email_sent:
#                             recovery_msg = (
#                                 f"✅ The stream has recovered at {recovery_timestamp}.\n"
#                                 f"❄️ Issue started at: {freeze_start_str}\n"
#                                 f"⏱️ Total freeze duration: {freeze_duration_str}"
#                             )
#                             log_freeze_event("✅ Stream recovered from freeze.", excel_file)
#                             send_email("✅ Stream Recovery Notice", recovery_msg)
#                         else:
#                             print("[INFO] Freeze recovery detected during cooldown — skipping recovery email.")
#
#                         freeze_logged = False
#                         freeze_email_sent = False
#                         freeze_start_time = None
#                         freeze_count = 0
#
#             prev_frame = frame
#             time.sleep(0.5)
#
#     # Start both threads
#     threading.Thread(target=audio_thread, daemon=True).start()
#     threading.Thread(target=video_thread, daemon=True).start()
#
#     try:
#         while not stop_event.is_set():
#             time.sleep(1)
#     except KeyboardInterrupt:
#         print("[INFO] Interrupted by user.")
#     finally:
#         print("[INFO] Stopping monitoring...")
#         try:
#             driver.quit()
#         except WebDriverException:
#             pass
#
#
# # -------------------- ENTRY POINT --------------------
# if __name__ == "__main__":
#     stop_event = threading.Event()
#     try:
#         main(
#             tester_url="https://www.radiantmediaplayer.com/stream-tester.html",
#             stream_url="https://cdn01khi.tamashaweb.com:8087/YlUHeDQb7a/city41News-abr/playlist.m3u8",
#             stop_event=stop_event
#         )
#     except KeyboardInterrupt:
#         print("[INFO] Caught keyboard interrupt. Exiting...")
#     finally:
#         stop_event.set()


# Updated Script


# Rewriting and saving the updated main.py again after reset
#
#
# # main.py

import os
import time
import threading
import smtplib
import winsound
import cv2
import numpy as np
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from mss import mss
from ctypes import POINTER, cast
import comtypes
from comtypes import CLSCTX_ALL
from pycaw.pycaw import AudioUtilities, IAudioMeterInformation

# -------------------- CONFIGURATION --------------------
SCREEN_REGION = {"top": 360, "left": 600, "width": 1000, "height": 560}
ALERT_SOUND = "alert.mp3"
EMAIL_ALERT = True
EMAIL_FROM = "expressdartt@gmail.com"
EMAIL_TO = "alibasti2021@gmail.com"
EMAIL_CC = "m.minds456@gmail.com"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USERNAME = "expressdartt@gmail.com"
SMTP_PASSWORD = "ztonyoyowxsdsarc"
ALERT_COOLDOWN_SECONDS = 300

# ==== GLOBAL STATE ====
last_video_alert_time = 0
last_audio_alert_time = 0
stop_event = threading.Event()
is_stream_frozen_flag = threading.Event()


# -------------------- UTILITIES --------------------
def play_alert():
    try:
        print("[AUDIO] Playing beep alert...")
        winsound.Beep(1000, 500)
    except Exception as e:
        print(f"[AUDIO] Beep failed: {e}")


def send_email(subject, body, attachment_path=None, theme="alert", stream_url=None):
    if not EMAIL_ALERT:
        return
    try:
        msg = MIMEMultipart("alternative")
        msg["From"] = EMAIL_FROM
        msg["To"] = EMAIL_TO
        msg["CC"] = EMAIL_CC
        msg["Subject"] = subject

        # Set color theme
        header_color = {
            "alert": "#ff4d4d",  # red
            "recovery": "#28a745",  # green
            "default": "#007bff"  # blue
        }.get(theme, "#007bff")

        # Replace newline characters in body
        html_friendly_body = body.replace("\n", "<br>")

        # Construct HTML content
        stream_url_html = f"<p><strong>📺 Stream URL:</strong><br><a href='{stream_url}'>{stream_url}</a></p>" if stream_url else ""
        html_body = f"""
        <html>
            <body style="font-family:Arial, sans-serif; background-color:#f9f9f9; padding:20px;">
                <div style="max-width:600px; margin:auto; border:1px solid #ddd; border-radius:8px; background-color:white;">
                    <div style="background-color:{header_color}; color:white; padding:15px 20px;
                    border-top-left-radius:8px; border-top-right-radius:8px;"> <h2 style="margin:0;">{subject}</h2>
                    </div>
                    <div style="padding:20px; color:#333;">
                        <p>{html_friendly_body}</p>
                        {stream_url_html}
                    </div>
                </div>
            </body>
        </html>
        """

        # Plain text fallback
        text_body = f"{subject}\n\n{body}"
        if stream_url:
            text_body += f"\n\nStream URL: {stream_url}"

        msg.attach(MIMEText(text_body, "plain"))
        msg.attach(MIMEText(html_body, "html"))

        if attachment_path:
            with open(attachment_path, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f"attachment; filename={os.path.basename(attachment_path)}")
                msg.attach(part)

        recipients = [EMAIL_TO] + EMAIL_CC.split(",")
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.sendmail(EMAIL_FROM, recipients, msg.as_string())
        print("[EMAIL] Themed email sent.")
    except Exception as e:
        print(f"[EMAIL] Failed to send themed email: {e}")


def initialize_excel_log():
    filename = f"freeze_log_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
    if not os.path.exists(filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Freeze Events"
        ws.append(["Timestamp", "Event Description"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        wb.save(filename)
    return filename


def log_freeze_event(message, excel_file):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{timestamp}] {message}")
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append([timestamp, message])
    wb.save(excel_file)


def is_stream_frozen(prev_frame, curr_frame, threshold=500):
    diff = cv2.absdiff(prev_frame, curr_frame)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    return cv2.countNonZero(gray) < threshold


def setup_browser(tester_url, stream_url):
    options = Options()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)
    driver.get(tester_url)
    time.sleep(3)
    try:
        input_field = driver.find_element(By.XPATH, "//*[@id='stream-url']")
        driver.execute_script("arguments[0].scrollIntoView(true);", input_field)
        time.sleep(1)
        input_field.clear()
        input_field.send_keys(stream_url)
        input_field.send_keys(Keys.ENTER)
        driver.find_element(By.XPATH, "//*[@id='rmp']/button").click()
    except Exception as e:
        print(f"[ERROR] Failed to enter stream URL: {e}")
    time.sleep(8)
    return driver


# -------------------- MAIN MONITORING FUNCTION --------------------
def main(tester_url, stream_url, stop_event):
    excel_file = initialize_excel_log()

    last_audio_alert_time = datetime.min
    last_video_alert_time = datetime.min
    audio_alert_suppressed = False
    video_alert_suppressed = False

    def should_send_audio_alert():
        nonlocal last_audio_alert_time, audio_alert_suppressed
        now = datetime.now()
        if (now - last_audio_alert_time).total_seconds() >= ALERT_COOLDOWN_SECONDS:
            last_audio_alert_time = now
            audio_alert_suppressed = False
            return True
        elif not audio_alert_suppressed:
            remaining = ALERT_COOLDOWN_SECONDS - (now - last_audio_alert_time).total_seconds()
            print(f"[AUDIO ALERT PAUSED] Next audio alert in {int(remaining)} seconds.")
            audio_alert_suppressed = True
        return False

    def should_send_video_alert():
        nonlocal last_video_alert_time, video_alert_suppressed
        now = datetime.now()
        if (now - last_video_alert_time).total_seconds() >= ALERT_COOLDOWN_SECONDS:
            last_video_alert_time = now
            video_alert_suppressed = False
            return True
        elif not video_alert_suppressed:
            remaining = ALERT_COOLDOWN_SECONDS - (now - last_video_alert_time).total_seconds()
            print(f"[VIDEO ALERT PAUSED] Next video alert in {int(remaining)} seconds.")
            video_alert_suppressed = True
        return False

    driver = setup_browser(tester_url, stream_url)

    def audio_thread():
        comtypes.CoInitialize()
        try:
            threshold = 0.02
            silence_duration = 2
            silent_start = None
            is_muted = False
            recovery_sent = False
            audio_email_sent = False

            devices = AudioUtilities.GetSpeakers()
            interface = devices.Activate(IAudioMeterInformation._iid_, CLSCTX_ALL, None)
            meter = cast(interface, POINTER(IAudioMeterInformation))

            while not stop_event.is_set():
                if is_stream_frozen_flag.is_set():
                    time.sleep(0.5)
                    continue

                volume = meter.GetPeakValue()
                if volume < threshold:
                    if silent_start is None:
                        silent_start = time.time()
                    elif time.time() - silent_start >= silence_duration and not is_muted:
                        if should_send_audio_alert():
                            send_email(
                                "🔇 Audio Muted Alert",
                                f"Audio has been muted for more than {silence_duration} seconds.\n🔗 Stream URL: {stream_url}",
                                theme="alert"
                            )
                            audio_email_sent = True
                        is_muted = True
                        recovery_sent = False
                else:
                    silent_start = None
                    if is_muted and not recovery_sent and audio_email_sent:
                        recovery_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        send_email(
                            "🔊 Audio Recovery Notice",
                            f"✅ Audio has resumed at {recovery_timestamp}.\n🔗 Stream URL: {stream_url}",
                            theme="recovery"
                        )
                        recovery_sent = True
                        is_muted = False
                        audio_email_sent = False

                time.sleep(0.1)
        finally:
            comtypes.CoUninitialize()

    def video_thread():
        sct = mss()
        prev_frame = None
        freeze_count = 0
        freeze_logged = False
        freeze_email_sent = False
        freeze_start_time = None

        print("[INFO] Monitoring stream...")

        while not stop_event.is_set():
            img = sct.grab(SCREEN_REGION)
            full_res_frame = np.array(img)
            full_res_frame = cv2.cvtColor(full_res_frame, cv2.COLOR_BGRA2BGR)
            frame = cv2.resize(full_res_frame, (600, 340))

            if prev_frame is not None:
                if is_stream_frozen(prev_frame, frame):
                    is_stream_frozen_flag.set()
                    freeze_count += 1
                    if not freeze_logged:
                        freeze_start_time = time.time()
                        log_freeze_event("Freeze detected...", excel_file)
                        freeze_logged = True
                        freeze_email_sent = False

                    if freeze_count >= 1:
                        if should_send_video_alert():
                            log_freeze_event("[ALERT] Stream appears to be frozen!", excel_file)
                            freeze_image_path = "freeze_screenshot.jpg"
                            cv2.imwrite(freeze_image_path, full_res_frame, [int(cv2.IMWRITE_JPEG_QUALITY), 80])
                            threading.Thread(target=play_alert, daemon=True).start()
                            send_email(
                                "🚫 Stream Freeze Detected",
                                f"The stream appears frozen.\n"
                                f"\n📺 Stream URL: {stream_url}",

                                freeze_image_path
                            )
                            freeze_email_sent = True
                        freeze_count = 0
                else:
                    if freeze_logged:
                        is_stream_frozen_flag.clear()
                        freeze_end_time = time.time()
                        freeze_duration = freeze_end_time - freeze_start_time
                        freeze_duration_str = time.strftime('%H:%M:%S', time.gmtime(freeze_duration))
                        freeze_start_str = datetime.fromtimestamp(freeze_start_time).strftime('%Y-%m-%d %H:%M:%S')
                        recovery_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                        if freeze_email_sent:
                            recovery_msg = (
                                f"✅ The stream has recovered at {recovery_timestamp}.\n"
                                f"❄️ Issue started at: {freeze_start_str}\n"
                                f"⏱️ Total freeze duration: {freeze_duration_str}\n"
                                f"🔗 Stream URL: {stream_url}"
                            )
                            log_freeze_event("✅ Stream recovered from freeze.", excel_file)
                            send_email("✅ Stream Recovery Notice", recovery_msg, theme="recovery")
                        freeze_logged = False
                        freeze_email_sent = False
                        freeze_start_time = None
                        freeze_count = 0

            prev_frame = frame
            time.sleep(0.5)

    threading.Thread(target=audio_thread, daemon=True).start()
    threading.Thread(target=video_thread, daemon=True).start()

    try:
        while not stop_event.is_set():
            time.sleep(1)
    except KeyboardInterrupt:
        print("[INFO] Interrupted by user.")
    finally:
        print("[INFO] Stopping monitoring...")
        stop_event.set()
        try:
            driver.quit()
        except:
            pass


if __name__ == "__main__":
    stop_event = threading.Event()
    try:
        main(
            tester_url="https://www.radiantmediaplayer.com/stream-tester.html",
            stream_url="https://cdn01khi.tamashaweb.com:8087/YlUHeDQb7a/city41News-abr/playlist.m3u8",
            stop_event=stop_event
        )
    except KeyboardInterrupt:
        print("[INFO] Caught keyboard interrupt. Exiting...")
    finally:
        stop_event.set()


